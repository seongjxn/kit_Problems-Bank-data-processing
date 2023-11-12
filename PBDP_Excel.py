# openpyxl Packages
from openpyxl import load_workbook as openpyxl_load_workbook
from openpyxl import Workbook as openpyxl_Workbook
from openpyxl.styles import Font as openpyxl_styles_Font


# Open Excel File
def open_excel_file(filename: str) -> openpyxl_Workbook:
    """Open an Excel file and return a workbook object."""
    wb = openpyxl_Workbook()
    wb = openpyxl_load_workbook(filename)
    return wb


# Create Excel File
def create_excel_file() -> openpyxl_Workbook:
    """Create an Excel file"""
    wb = openpyxl_Workbook()
    
    header = ['문제순서', '문제구분', '문제내용', '보기1', '보기2', '보기3', '보기4', '보기5', '정답']
    write_row_to_excel_file(wb, header, header=True)
    
    return wb


# Save Excel File
def save_excel_file(wb: openpyxl_Workbook, filename: str) -> bool:
    """Save an Excel file."""
    try:
        wb.save(filename)
        return True
    except:
        return False


# Write to Excel File
def write_row_to_excel_file(wb: openpyxl_Workbook, row: list, header=False) -> openpyxl_Workbook:
    """Write a row to an Excel file."""
    ws = wb.active
    ws.append(row)
    
    if header:
        header_row = ws[1]
        for cell in header_row:
            cell.font = openpyxl_styles_Font(bold=True, color= '000000FF')
    
    return wb


# Append Problem to Excel File
def append_problem_to_excel_file(wb: openpyxl_Workbook,problem: dict):
    """Append a problem to an Excel file."""
    problem_list = [problem['order'], '', problem['content'], problem['choice1'], problem['choice2'], problem['choice3'], problem['choice4'], '', problem['answer']]
    write_row_to_excel_file(wb, problem_list)



# Read from Excel File
def read_problems_from_excel_file(wb: openpyxl_Workbook):
    """Read rows from an Excel file."""
    ws = wb.active
    problems = []
    problem = {}
    for i, row in enumerate(ws.iter_rows()):
        if i % 3 == 0:
            problem['order'] = row[0].value
            problem['content'] = row[1].value
        if i % 3 == 1:
            problem['choice1'] = row[0].value
            problem['choice2'] = row[1].value
            problem['choice3'] = row[2].value
            problem['choice4'] = row[3].value
        if i % 3 == 2:
            problem['answer'] = row[1].value
            problems.append(problem)
            problem = {}
    return problems