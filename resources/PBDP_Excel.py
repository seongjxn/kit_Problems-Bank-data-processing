# openpyxl Packages
from openpyxl import load_workbook as openpyxl_load_workbook
from openpyxl import Workbook as openpyxl_Workbook
from openpyxl.styles import Font as openpyxl_styles_Font


# Open Excel File
def open_excel_file(filename: str) -> openpyxl_Workbook:
    """Open an Excel file and return a workbook object."""
    wb = openpyxl_Workbook()
    wb = openpyxl_load_workbook(filename)
    return wb


# Create Excel File
def create_excel_file() -> openpyxl_Workbook:
    """Create an Excel file"""
    wb = openpyxl_Workbook()
    ws = wb.active

    header = ['문제순서', '문제구분', '문제내용', '보기1', '보기2', '보기3', '보기4', '보기5', '정답']
    write_to_excel_file(wb, header, header=True)

    return wb


# Save Excel File
def save_excel_file(wb: openpyxl_Workbook, filename: str) -> bool:
    """Save an Excel file."""
    try:
        wb.save(filename)
        return True
    except:
        return False


# Write to Excel File
def write_row_to_excel_file(wb: openpyxl_Workbook, row: list, header=False):
    """Write a row to an Excel file."""
    ws = wb.active
    ws.append(row)

    if header:
        header_row = ws[1]
        for cell in header_row:
            cell.font = openpyxl_styles_Font(bold=True, color= '000000FF')


# Read from Excel File
def read_rows_from_excel_file()